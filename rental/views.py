import math
import os
import urllib.request
import urllib.parse
import json
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib import messages
from django.core.mail import send_mail
from django.core import signing
from django.conf import settings
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.contrib.sites.shortcuts import get_current_site
from django.contrib.gis.db.models.functions import Distance, Cast
from django.contrib.gis.db.models import GeometryField
from django.db.models import BinaryField
from django.contrib.gis.measure import D
from django.contrib.gis.geos import Point, Polygon
from django.core.paginator import Paginator
from django.utils import timezone
from django.utils.timezone import make_aware
from django.utils.encoding import force_bytes, force_str

from django.core.mail import send_mail
from django.conf import settings

from django.db.models import Count
from datetime import datetime
from functools import wraps

import openpyxl
from openpyxl.styles import Alignment
from .models import BookingHistory, Car, CarType, GPSLog, SafeZone, UserCustom, Station
from .gis_tools import calculate_stats, START_POINT_COORD

from django.db.models import Sum
from django.db.models.functions import TruncDate
from django.utils import timezone
from datetime import timedelta

from django.http import HttpResponse

# Điểm gốc mặc định (Bến Thành) - Kinh độ trước, Vĩ độ sau cho GIS
START_POINT_GIS = Point(106.698, 10.771, srid=4326)

# Cấu hình Vùng an toàn hình tròn
SAFE_CIRCLE_CENTER = [10.771, 106.698] # [lat, lng]
SAFE_RADIUS_METERS = 5000 # 5km

# Hàm bổ trợ kiểm tra quyền (Thay thế cho decorator của Django)
def is_admin(request):
    return request.session.get('role') == 'admin'

def is_logged_in(request):
    return 'user_id' in request.session

def admin_only(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if request.session.get('role') == 'admin':
            return view_func(request, *args, **kwargs)
        else:
            # Nếu không phải admin, điều hướng ra trang 403
            return render(request, 'rental/403.html', status=403)
    return _wrapped_view

# --- 1. HỆ THỐNG ĐĂNG NHẬP & QUẢN LÝ TÀI KHOẢN TỰ CHẾ ---

def custom_login(request):
    if request.method == 'POST':
        user_name = request.POST.get('username')
        pass_word = request.POST.get('password')
        
        try:
            # Truy vấn từ bảng UserCustom của bạn
            user = UserCustom.objects.get(username=user_name, password=pass_word)
            
            # Lưu thông tin vào session để duy trì đăng nhập
            request.session['user_id'] = user.id
            request.session['username'] = user.username
            request.session['role'] = user.role
            request.session['irl_name'] = user.IRL_name
            
            return redirect('home')
        except UserCustom.DoesNotExist:
            messages.error(request, "Tên đăng nhập hoặc mật khẩu không chính xác.")
            
    return render(request, 'registration/login.html')

def custom_logout(request):
    request.session.flush() # Xóa toàn bộ session
    return redirect('login')

# Trang quản lý tài khoản (Chỉ xem danh sách)
@admin_only
def account_management(request):
    if not is_admin(request):
        return render(request, 'rental/403.html', status=403)
    
    users = UserCustom.objects.all()
    return render(request, 'rental/account_management.html', {'users': users})

# API xử lý Thêm và Sửa tài khoản (Dùng cho Modal Popup)
def upsert_user(request):
    if request.method == 'POST' and is_admin(request):
        user_id = request.POST.get('user_id')
        username = request.POST.get('username')
        password = request.POST.get('password')
        irl_name = request.POST.get('IRL_name')
        role = request.POST.get('role')
        
        try:
            if user_id: # Trường hợp Sửa (Update)
                user = get_object_or_404(UserCustom, id=user_id)
                user.username = username
                if password: # Chỉ cập nhật mật khẩu nếu có nhập mới
                    user.password = password
                user.IRL_name = irl_name
                user.role = role
                user.save()
            else: # Trường hợp Thêm mới (Create)
                UserCustom.objects.create(
                    username=username, 
                    password=password, 
                    IRL_name=irl_name, 
                    role=role
                )
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Yêu cầu không hợp lệ'}, status=403)

# Xóa tài khoản
def delete_user(request, user_id):
    if not is_admin(request):
        return redirect('login')
    user = get_object_or_404(UserCustom, id=user_id)
    user.delete()
    messages.success(request, "Đã xóa tài khoản thành công.")
    return redirect('account_management')

from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
from django.template.loader import render_to_string
from django.core import signing  # Sử dụng signing để thay thế token generator
from django.contrib.sites.shortcuts import get_current_site
from .models import UserCustom

def register_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        email = request.POST.get('email')
        full_name = request.POST.get('full_name')

        # 1. Kiểm tra tồn tại
        if UserCustom.objects.filter(username=username).exists():
            messages.error(request, "Tên đăng nhập đã tồn tại.")
            return render(request, 'registration/register.html')

        if UserCustom.objects.filter(email=email).exists():
            messages.error(request, "Email này đã được sử dụng.")
            return render(request, 'registration/register.html')

        try:
            # 2. Tạo người dùng (is_active=False để khóa cho đến khi xác thực)
            user = UserCustom.objects.create(
                username=username,
                password=password, 
                email=email,
                IRL_name=full_name,
                role='guest',
                is_active=False
            )

            # 3. Tạo Token an toàn chứa ID người dùng
            # Token này tự động mã hóa bằng SECRET_KEY của Django
            token = signing.dumps({'user_id': user.pk})

            # 4. Gửi mail
            current_site = get_current_site(request)
            subject = 'Kích hoạt tài khoản của bạn'
            message = render_to_string('registration/account_activate.html', {
                'user': user,
                'domain': current_site.domain,
                'token': token,
            })

            send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [email])

            messages.warning(request, "Đăng ký thành công! Vui lòng kiểm tra Mailtrap để kích hoạt tài khoản.")
            return redirect('login')

        except Exception as e:
            messages.error(request, f"Lỗi hệ thống: {str(e)}")
            
    return render(request, 'registration/register.html')

def activate(request, token):
    try:
        # Giải mã token (hết hạn sau 2 ngày = 172800 giây)
        data = signing.loads(token, max_age=172800)
        user = UserCustom.objects.get(pk=data['user_id'])
    except (signing.SignatureExpired, signing.BadSignature, UserCustom.DoesNotExist):
        user = None

    if user is not None:
        user.is_active = True
        user.save()
        messages.success(request, "Tài khoản của bạn đã được kích hoạt thành công!")
        return redirect('login')
    else:
        messages.error(request, "Link kích hoạt không hợp lệ hoặc đã hết hạn.")
        return redirect('register')
# --- 2. CÁC VIEW GIAO DIỆN CHÍNH (KHÔNG THAY ĐỔI LOGIC) -----

def home(request):
    return render(request, 'rental/index.html')

def about(request):
    return render(request, 'rental/about.html')

@admin_only
def administration(request):
    return render(request, 'rental/administration.html')

def car_types_view(request):
    car_types = CarType.objects.all()
    return render(request, 'rental/car_types.html', {'car_types': car_types})

def car_detail_view(request, type_id):
    car_type = get_object_or_404(CarType, id=type_id)
    specific_cars = Car.objects.filter(car_type=car_type)
    return render(request, 'rental/car_detail.html', {
        'type': car_type,
        'cars': specific_cars
    })

def book_car(request):
    if request.method == 'POST':
        user_id = request.session.get('user_id')
        if not user_id:
            return JsonResponse({'status': 'error', 'message': 'Vui lòng đăng nhập'}, status=401)

        car_id = request.POST.get('car_id')    
        type_id = request.POST.get('car_type_id')
        station_id = request.POST.get('station_id') 
        try:
            # 1. Tìm xe phù hợp
            if car_id:
                car_to_book = get_object_or_404(Car, id=car_id, is_available=True)
            elif station_id:
                station = get_object_or_404(Station, id=station_id)
                car_to_book = Car.objects.filter(
                    car_type_id=type_id,
                    current_location__within=station.area,
                    is_available=True
                ).first()
            else:
                car_to_book = Car.objects.filter(
                    car_type_id=type_id,
                    is_available=True
                ).first()
                
            if not car_to_book:
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Hiện tại loại xe này đã hết xe sẵn sàng!'
                })
            # 2. Tối ưu: Nếu đặt từ danh sách chung, tự tìm trạm mà xe đang đỗ
            # (Giúp lịch sử luôn có thông tin trạm bắt đầu)
            actual_pickup_station = None
            if station_id:
                actual_pickup_station = station
            else:
                actual_pickup_station = Station.objects.filter(
                    area__contains=car_to_book.current_location
                ).first()
            # 3. Tạo bản ghi lịch sử
            BookingHistory.objects.create(
                user=UserCustom.objects.get(id=user_id),
                car=car_to_book,
                pickup_station=actual_pickup_station,
                status='ongoing'
            )
            # 4. Cập nhật trạng thái xe
            car_to_book.is_available = False
            car_to_book.save()
            # 5. Cập nhật trạng thái User (Nếu bạn vẫn dùng bảng UserCustom để quản lý đơn hàng)
            user = UserCustom.objects.get(id=user_id)
            full_name = request.POST.get('full_name')
            phone = request.POST.get('phone_number')
            email = request.POST.get('email')
            book_date = request.POST.get('book_date') # Ngày nhận xe (YYYY-MM-DD)
            book_time = request.POST.get('book_time') # Giờ nhận xe (HH:MM)

            # Cập nhật thông tin cá nhân
            user.IRL_name = full_name
            user.phone_number = phone
            user.email = email
            
            # Ghi nhận thời điểm thực hiện đăng ký đơn
            user.book_day = timezone.now() 

            # Xử lý gộp Ngày + Giờ thành pickup_day
            if book_date and book_time:
                try:
                    pickup_str = f"{book_date} {book_time}"
                    # Chuyển string thành object datetime
                    naive_dt = datetime.strptime(pickup_str, '%Y-%m-%d %H:%M')
                    # Chuyển thành aware datetime (có múi giờ) để lưu vào database
                    user.pickup_day = make_aware(naive_dt)
                except ValueError:
                    pass

            # SỬA LỖI: Luôn để trạng thái là pending khi mới đặt xong
            user.order_status = 'pending' 
            user.save()

            return JsonResponse({'status': 'success', 'message': 'Đặt xe thành công!'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@admin_only        
def dashboard(request):
    if not is_logged_in(request):
        return redirect('login')

    cars = Car.objects.all()
    for car in cars:
        stats = calculate_stats(car)
        car.total_km = stats['total_km']
        car.total_fee = stats['total_fee']
        car.base_fee = stats.get('base_fee', 0)
        car.extra_fee = stats.get('extra_fee', 0)
        car.penalty_fee = stats.get('penalty_fee', 0)

    safe_zones = SafeZone.objects.all()
    zones_data = []
    for zone in safe_zones:
        centroid = zone.area.centroid
        zones_data.append({
            'name': zone.name,
            'coords': [centroid.y, centroid.x],
        })
    
    return render(request, 'rental/dashboard.html', {
        'cars': cars,
        'zones_data': zones_data,
        'START_POINT': SAFE_CIRCLE_CENTER,
        'SAFE_RADIUS': SAFE_RADIUS_METERS
    })

@admin_only
def management(request):
    if not is_admin(request):
        messages.error(request, "Bạn cần quyền Admin để vào trang giả lập điều khiển.")
        return redirect('login')

    cars = Car.objects.all()
    safe_zones = SafeZone.objects.all()
    zones_data = []
    for zone in safe_zones:
        coords = [[p[1], p[0]] for p in zone.area[0]]
        zones_data.append({'name': zone.name, 'coords': coords})
        
    return render(request, 'rental/management.html', {
        'cars': cars, 
        'zones_data': zones_data,
        'START_POINT': SAFE_CIRCLE_CENTER,
        'SAFE_RADIUS': SAFE_RADIUS_METERS
    })

# --- 3. API XỬ LÝ LOGIC GIS (KHÔNG THAY ĐỔI LOGIC) ---

def simulate_movement(request):
    if not is_logged_in(request):
        return JsonResponse({'status': 'error', 'message': 'Unauthorized'}, status=401)
    try:
        car_id = request.GET.get('car_id')
        target_lat = float(request.GET.get('lat'))
        target_lon = float(request.GET.get('lon'))
        
        car = get_object_or_404(Car, id=car_id)
        current_pos = car.current_location
        old_stats = calculate_stats(car)

        new_lon = target_lon
        new_lat = target_lat

        car.current_location = Point(new_lon, new_lat, srid=4326)
        car.save()

        car.current_location = Point(new_lon, new_lat, srid=4326)
        car.save()

        GPSLog.objects.create(car=car, location=car.current_location)
        new_stats = calculate_stats(car)

        return JsonResponse({
            'status': 'success',
            'new_lat': new_lat,
            'new_lng': new_lon,
            'old_km': old_stats['total_km'],
            'new_km': new_stats['total_km'],
            'old_fee': old_stats['total_fee'],
            'new_fee': new_stats['total_fee']
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

def reset_data(request):
    if not is_admin(request):
        return JsonResponse({'status': 'error', 'message': 'Admin Only'}, status=403)
    GPSLog.objects.all().delete()
    Car.objects.all().update(current_location=START_POINT_GIS)
    return JsonResponse({'status': 'success'})

def update_order_status(request, user_id, status):
    if not is_admin(request):
        return redirect('login')
        
    user = get_object_or_404(UserCustom, id=user_id)
    # Cập nhật trạng thái theo tham số truyền vào từ URL (confirmed hoặc cancelled)
    user.order_status = status
    user.save()
    
    messages.success(request, f"Đã cập nhật trạng thái đơn hàng cho {user.username} thành: {user.get_order_status_display()}")
    return redirect('account_management')

def delete_order(request, user_id):
    if not is_admin(request):
        return redirect('login')
    
    user = get_object_or_404(UserCustom, id=user_id)
    # Reset toàn bộ thông tin đơn hàng
    user.book_day = None
    user.pickup_day = None # Xóa ngày hẹn nhận xe
    user.email = None
    user.phone_number = None
    user.order_status = 'pending' 
    user.IRL_name = "" # Reset tên thật về rỗng
    user.save()
    
    messages.success(request, f"Đã xóa hóa đơn và làm sạch dữ liệu cho {user.username}")
    return redirect('account_management')

def checkout_car(request):
    if not is_logged_in(request):
        return JsonResponse({'status': 'error'}, status=401)
        
    if request.method == 'POST':
        car_id = request.POST.get('car_id')
        car = get_object_or_404(Car, id=car_id)

        # 1. Tìm chuyến đi đang diễn ra (ongoing) của xe này [cite: 136]
        history = BookingHistory.objects.filter(car=car, status='ongoing').last()
        
        if history:
            # 2. Tính toán stats tại thời điểm nhấn nút thanh toán 
            stats = calculate_stats(car)
            
            # 3. Chốt dữ liệu chuyến đi vào lịch sử 
            history.status = 'completed'
            history.end_time = timezone.now()
            history.total_distance = stats['total_km']
            history.total_fee = stats['total_fee']
            # Giữ nguyên logic cũ: trạm trả mặc định là trạm lấy xe 
            history.return_station = history.pickup_station 
            history.save()

        # 4. Xóa lịch sử GPS cũ của xe 
        GPSLog.objects.filter(car=car).delete()

        # 5. Reset trạng thái xe về sẵn sàng tại điểm gốc 
        car.current_location = START_POINT_GIS
        car.is_available = True 
        car.save()

        return JsonResponse({'status': 'success'})

def get_cars_in_zone(request):
    try:
        bounds = request.GET.get('bounds', '').split(',')
        if len(bounds) == 4:
            w, s, e, n = map(float, bounds)
            cars = Car.objects.filter(
                current_location__x__gte=w, current_location__x__lte=e,
                current_location__y__gte=s, current_location__y__lte=n
            )
            data = [{'plate': c.license_plate, 'type': c.car_type.name} for c in cars]
            return JsonResponse({'cars': data})
        return JsonResponse({'cars': []})
    except Exception as err:
        return JsonResponse({'error': str(err)}, status=400)

# --- 4. QUẢN LÝ THÊM/XÓA XE (DÙNG CHO TRANG ADMIN RIÊNG) ---

def get_addCar(request):
    if not is_admin(request):
        return redirect('login')
    car_list = CarType.objects.all()
    return render(request, 'carAdd.html', {'car_list': car_list})

@admin_only
def add_car(request):
    if not is_admin(request):
        return redirect('login')
    if request.method == 'POST':
        type_id = request.POST.get('car_type_id')
        brand = request.POST.get('brand_name')
        plate = request.POST.get('license_plate')
        price = request.POST.get('base_price_per_km')
        image = request.FILES.get('car_image')
        
        car_type = get_object_or_404(CarType, id=type_id)
        if price:
            car_type.base_price_per_km = price
            car_type.save()
            
        Car.objects.create(
            car_type=car_type,
            brand_name=brand,
            license_plate=plate,
            image=image,
            current_location=START_POINT_GIS
        )
        return redirect('car_types')
    return redirect('add_car')

def update_car_short_description(request):
    if not is_admin(request):
        return redirect('login')
    if request.method != 'POST':
        return redirect('car_types')
    car_id = request.POST.get('car_id')
    type_id = request.POST.get('car_type_id')
    text = (request.POST.get('short_description') or '').strip()
    if len(text) > 2000:
        text = text[:2000]
    car = get_object_or_404(Car, id=car_id, car_type_id=type_id)
    car.short_description = text
    car.save()
    return redirect('car_detail', type_id=type_id)


def delete_car(request, id):
    if not is_admin(request):
        return redirect('login')
    car = get_object_or_404(Car, id=id)
    if car.image:
        try:
            if os.path.isfile(car.image.path):
                os.remove(car.image.path)
        except:
            pass
    car.delete()
    return redirect('car_types')

def station_detail(request, station_id):
    station = get_object_or_404(Station, id=station_id)
    
    # Lấy danh sách xe đang ở trong vùng của trạm và đang sẵn sàng
    cars_in_station = Car.objects.filter(
        current_location__within=station.area,
        is_available=True
    )
    
    # Logic Group by: Đếm số lượng xe theo từng Hãng và Loại xe
    inventory = cars_in_station.values('brand_name', 'car_type__name').annotate(
        total=Count('id') # Đếm số ID xe trùng hãng
    ).order_by('brand_name')

    current_count = cars_in_station.count()
    available_slots = station.capacity - current_count
    
    context = {
        'station': station,
        'inventory': inventory,
        'current_count': current_count,
        'available_slots': max(0, available_slots),
        'is_full': current_count >= station.capacity,
    }
    return render(request, 'rental/station_detail.html', context)

def add_station(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        lat = float(request.POST.get('lat'))
        lon = float(request.POST.get('lon'))
        capacity = request.POST.get('capacity', 10)

        if capacity < 0:
            messages.error(request, "Sức chứa không được là số âm!")
            return redirect('station_list')
        
        # 1. Tạo Point từ tọa độ (Kinh độ trước, Vĩ độ sau cho GIS)
        pnt = Point(lon, lat, srid=4326)
        
        # 2. Tạo vùng Polygon nhỏ quanh điểm đó (khoảng 50m) 
        # Trong thực tế 0.0005 độ xấp xỉ 50m
        station_area = pnt.buffer(0.0005) 
        
        # 3. Lưu vào Database
        Station.objects.create(
            name=name,
            area=station_area,
            capacity=capacity
        )
        
        messages.success(request, f"Đã thêm trạm {name} thành công!")
        return redirect('station_list')
        
    return redirect('station_list')

def validate_return(car, return_location):
    # 1. Tìm trạm chứa tọa độ trả xe
    station = Station.objects.filter(area__contains=return_location).first()
    
    if not station:
        return False, "Bạn đang đứng ngoài khu vực trạm xe. Vui lòng di chuyển vào vùng quy định."

    # 2. Kiểm tra sức chứa của trạm đó
    current_total = station.current_car_count()
    if current_total >= station.capacity:
        return False, f"Trạm {station.name} đã hết chỗ đỗ ({current_total}/{station.capacity})."

    # 3. Cập nhật vị trí VÀ cho phép xe có thể được đặt tiếp
    car.current_location = return_location
    car.is_available = True  # THÊM DÒNG NÀY: Để xe trở lại trạng thái sẵn sàng
    car.save()
    
    return True, f"Trả xe thành công vào trạm {station.name}!"

def api_return_car(request):
    car_id = request.GET.get('car_id')
    lat = float(request.GET.get('lat'))
    lon = float(request.GET.get('lon'))
    
    car = Car.objects.get(id=car_id)
    return_pos = Point(lon, lat) # Lưu ý: Point nhận (lon, lat)
    
    # Gọi hàm xử lý đã viết ở trên
    success, message = validate_return(car, return_pos)
    
    if success:
        return JsonResponse({
            'status': 'success',
            'message': message,
            'redirect_url': f'/station/{Station.objects.filter(area__contains=return_pos).first().id}/'
        })
    else:
        return JsonResponse({'status': 'error', 'message': message})
# rental/views.py
from django.contrib.gis.measure import D

def station_list(request):
    query = request.GET.get('q', '')
    lat = request.GET.get('lat')
    lon = request.GET.get('lon')
    radius = request.GET.get('radius')
    
    # Khởi tạo query ban đầu
    stations_list = Station.objects.all().order_by('name')

    # Xử lý tìm kiếm theo tên
    if query:
        stations_list = stations_list.filter(name__icontains=query)

    # Xử lý tìm kiếm theo vị trí (GIS)
    if lat and lon and radius:
        try:
            user_point = Point(float(lon), float(lat), srid=4326)
            # Dùng dwithin để lọc các trạm có "area" nằm trong bán kính
            stations_list = stations_list.filter(area__dwithin=(user_point, float(radius)))
        except (ValueError, TypeError):
            pass

    # Phân trang
    paginator = Paginator(stations_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'rental/station_list.html', {
        'page_obj': page_obj,
        'stations': page_obj.object_list,
        'query': query
    })

@admin_only
def delete_station(request, station_id):
    station = get_object_or_404(Station, id=station_id)
    name = station.name
    station.delete()
    messages.success(request, f"Đã xóa trạm {name} thành công.")
    return redirect('station_list')

from django.contrib.gis.db.models.functions import Distance
from django.contrib.gis.measure import D

def get_stations_nearby(request):
    try:
        # 1. Lấy dữ liệu và ép kiểu
        lat_raw = request.GET.get('lat')
        lon_raw = request.GET.get('lon')
        radius_raw = request.GET.get('radius', 2000)

        if not lat_raw or not lon_raw:
            return JsonResponse({'error': 'Missing coordinates'}, status=400)

        lat = float(lat_raw)
        lon = float(lon_raw)
        radius = float(radius_raw)

        # 2. Tạo điểm Point (Kinh độ trước, Vĩ độ sau)
        user_location = Point(lon, lat, srid=4326)

        # 3. Truy vấn (Sử dụng trường 'area' kiểu PolygonField của bạn)
        # area__distance_lte: tìm các đối tượng có khoảng cách nhỏ hơn hoặc bằng radius
        stations = Station.objects.filter(
            area__distance_lte=(user_location, D(m=radius))
        ).annotate(
            distance=Distance('area', user_location)
        ).order_by('distance')

        # 4. Format kết quả trả về
        results = []
        for st in stations:
            results.append({
                'id': st.id,
                'name': st.name,
                'lat': st.area.centroid.y, # Vĩ độ để Leaflet hiển thị
                'lng': st.area.centroid.x, # Kinh độ để Leaflet hiển thị
                'distance': st.distance.m   # Khoảng cách tính bằng mét
            })

        return JsonResponse({'stations': results})

    except Exception as e:
        # Trả về nội dung lỗi cụ thể để bạn thấy được trong tab Response của trình duyệt
        return JsonResponse({'error': str(e)}, status=400)
    
# API để thêm trạm mới (Sử dụng GeoDjango để tạo Polygon từ Tâm + Bán kính)
def booking_history_list(request):
    user_id = request.session.get('user_id')
    role = request.session.get('role')
    # Nếu là Admin: Lấy tất cả lịch sử của mọi người dùng
    if role == 'admin':
        histories = BookingHistory.objects.all().order_by('-start_time')
    # Nếu là Guest: Chỉ lấy lịch sử của chính mình
    else:
        histories = BookingHistory.objects.filter(user_id=user_id).order_by('-start_time')
    for h in histories:
        if h.status == 'ongoing' and h.car:
            stats = calculate_stats(h.car)
            h.display_km = stats['total_km']
            h.display_fee = stats['total_fee']
            h.is_violated = stats['violated']
        else:
            h.display_km = h.total_distance
            h.display_fee = h.total_fee
            h.is_violated = False
    return render(request, 'rental/booking_history.html', {'histories': histories})

def return_car(request, history_id):
    if request.method == 'POST':
        history = get_object_or_404(BookingHistory, id=history_id)
        current_user_id = request.session.get('user_id')
        # KIỂM TRA: Chỉ cho phép trả xe nếu người đang đăng nhập là người đã đặt xe
        if history.user_id != current_user_id:
            from django.contrib import messages
            messages.error(request, "Bạn không có quyền trả xe của người khác!")
            return redirect('booking_history')
        if history.status == 'ongoing':
            stats = calculate_stats(history.car)
            # 1. Chốt dữ liệu chuyến đi cũ
            history.status = 'completed'
            history.end_time = timezone.now()
            history.total_distance = stats['total_km']
            history.total_fee = stats['total_fee']
            station = history.pickup_station
            history.return_station = station
            history.save()
            if history.car:
                # 2. XÓA LỊCH SỬ GPS CŨ
                GPSLog.objects.filter(car=history.car).delete()
                # 3. Đưa xe về trạng thái sẵn sàng tại tâm của trạm
                history.car.is_available = True
                history.car.current_location = station.area.centroid 
                history.car.save()
    return redirect('booking_history')

def get_daily_revenue_data(start_date, end_date):
    bookings = BookingHistory.objects.filter(
        status='completed',
        end_time__date__range=[start_date, end_date]
    ).select_related('car')

    daily_stats = {}
    for b in bookings:
        date = b.end_time.date()
        if date not in daily_stats:
            daily_stats[date] = {'total': 0, 'cars': set()}
        daily_stats[date]['total'] += float(b.total_fee)
        # Lấy tên xe (giả sử trường là car.brand hoặc car.name)
        car_name = b.car.brand if hasattr(b.car, 'brand') else str(b.car)
        daily_stats[date]['cars'].add(car_name)

    # Chuyển thành list và sắp xếp theo ngày
    results = []
    for date in sorted(daily_stats.keys()):
        results.append({
            'date': date,
            'total': daily_stats[date]['total'],
            'cars': ", ".join(daily_stats[date]['cars'])
        })
    return results

from django.db.models import Sum
from django.db.models.functions import TruncDate
from django.utils import timezone
from datetime import timedelta, datetime
from django.shortcuts import render
# Đảm bảo bạn đã import các model cần thiết
# from .models import BookingHistory 

@admin_only # Hoặc @admin_only tùy theo decorator bạn đang dùng
def revenue_statistics(request):
    # 1. Lấy tham số lọc từ URL
    filter_type = request.GET.get('filter', '7days')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    today = timezone.now().date()
    
    # 2. Thiết lập khoảng thời gian mặc định
    if filter_type == 'today':
        start_date = today
        end_date = today
    elif filter_type == '30days':
        start_date = today - timedelta(days=30)
        end_date = today
    elif filter_type == 'custom' and start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = today - timedelta(days=7)
            end_date = today
    else: # Mặc định 7 ngày
        start_date = today - timedelta(days=7)
        end_date = today

    # 3. Truy vấn dữ liệu từ database
    # Sử dụng select_related('car') để tối ưu hiệu năng khi lấy tên xe
    bookings = BookingHistory.objects.filter(
        status='completed',
        end_time__date__range=[start_date, end_date]
    ).select_related('car')

    # 4. Xử lý gộp dữ liệu theo từng ngày
    daily_stats = {}
    
    for b in bookings:
        # Lấy ngày từ end_time
        date_key = b.end_time.date()
        
        if date_key not in daily_stats:
            daily_stats[date_key] = {
                'total': 0,
                'cars': set() # Dùng set để tự động loại bỏ các xe trùng tên trong cùng 1 ngày
            }
        
        # Cộng dồn doanh thu
        daily_stats[date_key]['total'] += float(b.total_fee)
        
        # Lấy tên hãng xe hoặc loại xe (tùy thuộc vào trường bạn đặt trong model Car)
        # Ví dụ ở đây tôi lấy car.brand
        car_name = b.car.brand if hasattr(b.car, 'brand') else str(b.car)
        daily_stats[date_key]['cars'].add(car_name)

    # 5. Chuyển đổi dữ liệu sang dạng danh sách và sắp xếp theo ngày
    processed_data = []
    total_revenue = 0
    
    for date in sorted(daily_stats.keys()):
        total_revenue += daily_stats[date]['total']
        processed_data.append({
            'date': date,
            'total': daily_stats[date]['total'],
            'cars': ", ".join(daily_stats[date]['cars']) # Chuyển set xe thành chuỗi: "Toyota, Honda"
        })

    # 6. Trả về context cho template
    context = {
        'total_revenue': total_revenue,
        'processed_data': processed_data, # Dữ liệu đã gộp để vẽ biểu đồ và hiển thị bảng
        'filter_type': filter_type,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
    }
    
    return render(request, 'rental/revenue_statistics.html', context)

@admin_only
def export_revenue_excel(request):
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()

    # Sử dụng select_related('car') để tối ưu truy vấn thông tin xe
    bookings = BookingHistory.objects.filter(
        status='completed',
        end_time__date__range=[start_date, end_date]
    ).select_related('car')

    # 1. Logic gộp dữ liệu
    daily_report = {}
    for b in bookings:
        day = b.end_time.date()
        if day not in daily_report:
            daily_report[day] = {'total': 0, 'car_info': set()}
        
        daily_report[day]['total'] += float(b.total_fee)
        
        # Lấy biển số và hãng xe từ model Car
        lp = b.car.license_plate
        brand = b.car.brand_name if b.car.brand_name else "Không rõ hãng"
        
        # Tạo chuỗi định dạng "Biển số - Hãng xe"
        info_str = f"{lp} - {brand}"
        daily_report[day]['car_info'].add(info_str)

    # 2. Tạo file Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Thống kê doanh thu"
    
    headers = ['Ngày', 'Các loại xe đã đi', 'Tổng doanh thu ngày (VNĐ)']
    ws.append(headers)

    # Định dạng độ rộng cột cho dễ nhìn
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40 # Cột danh sách xe cần rộng hơn
    ws.column_dimensions['C'].width = 25

    # 3. Ghi dữ liệu đã gộp
    for day in sorted(daily_report.keys()):
        # Nối các xe bằng dấu phẩy và ký tự xuống dòng (\n)
        cars_formatted_str = ",\n".join(sorted(daily_report[day]['car_info']))
        
        # Ghi dòng dữ liệu
        row_data = [day.strftime('%d/%m/%Y'), cars_formatted_str, daily_report[day]['total']]
        ws.append(row_data)
        
        # Lấy ô (cell) vừa mới ghi ở cột B để thiết lập wrap_text (xuống dòng)
        current_row = ws.max_row
        cell = ws.cell(row=current_row, column=2)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=DoanhThu_{start_date_str}_To_{end_date_str}.xlsx'
    wb.save(response)
    return response

def error_404_view(request, exception=None):
    return render(request, 'rental/404.html', status=404)

# Hàm xử lý lỗi 405 (Quyền truy cập)
def error_403_view(request):
    return render(request, 'rental/403.html', status=403)

# Helper để kiểm tra quyền Admin
def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        # Kiểm tra đăng nhập [cite: 53]
        if not request.session.get('user_id'):
            return redirect('login')
        
        # Lấy thông tin user từ session và kiểm tra role 
        from .models import UserCustom
        user = UserCustom.objects.get(id=request.session['user_id'])
        if user.role != 'admin':
            return redirect('error_403') # Chuyển hướng đến trang 403 nếu không phải admin
            
        return view_func(request, *args, **kwargs)
    return wrapper